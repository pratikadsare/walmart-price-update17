
import io
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Walmart Price Update Tool (Bulk + Status)", layout="wide")
st.title("Walmart Price Update Tool (Bulk + Status)")

TEMPLATE_PATH = Path("templates/walmart_price_template.xlsx")

DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1jzEwuQY_1RAF296YmCAIxiu5ueznbBgx2nP5Rc_Yy2Y/edit?usp=sharing"

GSHEET_SKU_COL = "SKU"
GSHEET_STATUS_COL = "Publish Status"
GSHEET_PRICE_COL = "Price"

START_ROW = 7
COL_SKU = "D"
COLS_PRICE = ["E", "F", "G"]

MAX_ROWS = 1000

def extract_sheet_id(sheet_url: str) -> str:
    try:
        parts = sheet_url.split("/d/")
        if len(parts) < 2:
            return ""
        tail = parts[1]
        sheet_id = tail.split("/")[0]
        return sheet_id.strip()
    except Exception:
        return ""

def build_csv_export_url(sheet_url: str) -> str:
    sheet_id = extract_sheet_id(sheet_url)
    if not sheet_id:
        return ""
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

@st.cache_data(ttl=1800)
def load_status_sheet(csv_url: str) -> pd.DataFrame:
    return pd.read_csv(csv_url)

def normalize_sku(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s

def is_unpublished(status_val) -> bool:
    # Any status containing 'unpublished' (case-insensitive) is treated as unpublished.
    if status_val is None:
        return False
    s = str(status_val).strip().lower()
    return "unpublished" in s

def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return ""
    name = re.sub(r"[^\w\- ]+", "", name)
    name = re.sub(r"\s+", "_", name)
    return name

def clean_price_series(price_series: pd.Series) -> pd.Series:
    raw = price_series.copy()
    raw = raw.where(~raw.isna(), "")
    raw = raw.astype(str).str.strip()
    raw = raw.str.replace(",", "", regex=False)
    raw = raw.str.replace("₹", "", regex=False)
    raw = raw.str.replace("$", "", regex=False)
    return pd.to_numeric(raw, errors="coerce")

def fill_price_template(df: pd.DataFrame):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    max_rows_to_clear = max(ws.max_row, START_ROW + len(df) + 50)
    for r in range(START_ROW, max_rows_to_clear + 1):
        ws[f"{COL_SKU}{r}"].value = None
        for c in COLS_PRICE:
            ws[f"{c}{r}"].value = None

    for i, row in enumerate(df.itertuples(index=False), start=0):
        r = START_ROW + i
        sku = str(row.SKU).strip()
        price = float(row.New_Price)
        ws[f"{COL_SKU}{r}"].value = sku
        for c in COLS_PRICE:
            ws[f"{c}{r}"].value = price

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def validate_for_download(df: pd.DataFrame):
    # Hard Fail rules: blank SKU, invalid new price, duplicate SKU, SKU Not Found.
    # Soft Fail: Unpublished SKU -> needs confirmation checkbox.
    hard_errors = []
    if df is None or df.empty:
        return ["No rows found."], [], [], pd.DataFrame()

    d = df.copy()
    d["SKU"] = d["SKU"].apply(normalize_sku)
    d["New_Price_num"] = clean_price_series(d["New Price"])

    nonblank = d[(d["SKU"] != "") | (~d["New_Price_num"].isna())].copy()
    if nonblank.empty:
        return ["All rows are blank."], [], [], pd.DataFrame()

    if (nonblank["SKU"] == "").any():
        hard_errors.append("Some SKU values are blank.")
    if nonblank["New_Price_num"].isna().any():
        hard_errors.append("Some New Price values are blank or not a number.")
    if (nonblank["New_Price_num"] <= 0).any():
        hard_errors.append("Some New Price values are 0 or negative.")

    not_found_mask = nonblank["Publish Status"].astype(str).str.strip().eq("SKU Not Found")
    not_found_skus = nonblank.loc[not_found_mask, "SKU"].tolist()
    if not_found_skus:
        hard_errors.append(f"SKU Not Found on Walmart: {len(not_found_skus)}")

    dupes = nonblank[nonblank["SKU"] != ""].loc[nonblank["SKU"].duplicated(keep=False), "SKU"].unique().tolist()
    if dupes:
        hard_errors.append(f"Duplicate SKU found: {len(dupes)}")

    unpublished_mask = nonblank["Publish Status"].apply(is_unpublished)
    unpublished_skus = nonblank.loc[unpublished_mask & (~not_found_mask), "SKU"].tolist()

    writable = nonblank[
        (nonblank["SKU"] != "")
        & (~nonblank["New_Price_num"].isna())
        & (nonblank["New_Price_num"] > 0)
        & (~not_found_mask)
    ].copy()

    writable_out = writable[["SKU", "New_Price_num"]].copy()
    writable_out.columns = ["SKU", "New_Price"]
    return hard_errors, not_found_skus, unpublished_skus, writable_out

def apply_status_lookup(input_df: pd.DataFrame, status_df: pd.DataFrame):
    out = input_df.copy()
    out["SKU"] = out["SKU"].apply(normalize_sku)

    required = {GSHEET_SKU_COL, GSHEET_STATUS_COL, GSHEET_PRICE_COL}
    if not required.issubset(set(status_df.columns)):
        raise ValueError(f"Google Sheet must have columns: {', '.join(required)}")

    status_df = status_df.copy()
    status_df[GSHEET_SKU_COL] = status_df[GSHEET_SKU_COL].apply(normalize_sku)
    status_df = status_df[status_df[GSHEET_SKU_COL] != ""].drop_duplicates(subset=[GSHEET_SKU_COL], keep="last")

    status_map = dict(zip(status_df[GSHEET_SKU_COL], status_df[GSHEET_STATUS_COL].astype(str)))
    price_map = dict(zip(status_df[GSHEET_SKU_COL], status_df[GSHEET_PRICE_COL]))

    def get_status(sku: str) -> str:
        if not sku:
            return ""
        return str(status_map.get(sku, "SKU Not Found"))

    def get_curr_price(sku: str):
        if not sku:
            return ""
        val = price_map.get(sku, "")
        if pd.isna(val):
            return ""
        return val

    out["Publish Status"] = out["SKU"].apply(get_status)
    out["Current Price"] = out["SKU"].apply(get_curr_price)
    return out

with st.sidebar:
    st.header("Settings")
    sheet_url = st.text_input("Google Sheet link", value=DEFAULT_SHEET_URL)
    csv_url = build_csv_export_url(sheet_url)
    st.caption("Sheet must be shared as: Anyone with the link → Viewer")

    st.divider()
    st.header("Status")
    if not TEMPLATE_PATH.exists():
        st.error("Template missing")
        st.write("Add file at: templates/walmart_price_template.xlsx")
    else:
        st.success("Template found")

    if not csv_url:
        st.error("Invalid Google Sheet link")
    else:
        st.write("CSV source ready")

st.markdown("### 1) Choose number of rows")
row_count = st.number_input("Rows", min_value=1, max_value=MAX_ROWS, value=10, step=1)
st.caption("Tip: Copy 2 columns (SKU and New Price) from Excel and paste directly into the table.")

def empty_table(n: int) -> pd.DataFrame:
    return pd.DataFrame({"SKU": [""] * n, "New Price": [""] * n, "Publish Status": [""] * n, "Current Price": [""] * n})

if "table_df" not in st.session_state:
    st.session_state.table_df = empty_table(int(row_count))
else:
    current = st.session_state.table_df
    rc = int(row_count)
    if len(current) < rc:
        st.session_state.table_df = pd.concat([current, empty_table(rc - len(current))], ignore_index=True)
    elif len(current) > rc:
        st.session_state.table_df = current.iloc[:rc].reset_index(drop=True)

st.markdown("### 2) Paste SKU and New Price")
col_left, col_right = st.columns([3, 1])

with col_left:
    edited = st.data_editor(
        st.session_state.table_df,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "SKU": st.column_config.TextColumn("SKU"),
            "New Price": st.column_config.TextColumn("New Price"),
            "Publish Status": st.column_config.TextColumn("Publish Status"),
            "Current Price": st.column_config.TextColumn("Current Price"),
        },
        disabled=["Publish Status", "Current Price"],
        hide_index=True,
        key="editor",
    )
    st.session_state.table_df = edited

with col_right:
    st.markdown("### Quick info")
    tmp = edited.copy()
    tmp["SKU"] = tmp["SKU"].apply(normalize_sku)
    tmp["New_Price_num"] = clean_price_series(tmp["New Price"])
    tmp_nonblank = tmp[(tmp["SKU"] != "") | (~tmp["New_Price_num"].isna())].copy()
    st.metric("Rows filled", int(len(tmp_nonblank)))

    nf_count = int((tmp_nonblank["Publish Status"].astype(str).str.strip() == "SKU Not Found").sum())
    unpub_count = int(tmp_nonblank["Publish Status"].apply(is_unpublished).sum())

    if nf_count > 0:
        st.error(f"{nf_count} SKU Not Found")
    elif unpub_count > 0:
        st.warning(f"{unpub_count} Unpublished SKU")
    else:
        st.success("All OK")

st.divider()

c1, c2, c3 = st.columns([1, 1, 2])
with c1:
    if st.button("Clear table"):
        st.session_state.table_df = empty_table(int(row_count))
        st.rerun()
with c2:
    refresh = st.button("Refresh Status")
with c3:
    st.caption("Click **Refresh Status** after pasting SKU + New Price.")

if refresh:
    if not csv_url:
        st.error("Invalid Google Sheet link in sidebar.")
    else:
        try:
            status_df = load_status_sheet(csv_url)
            st.session_state.table_df = apply_status_lookup(st.session_state.table_df, status_df)
            st.success("Status updated from Google Sheet.")
            st.rerun()
        except Exception as e:
            st.error(f"Failed to read Google Sheet. Make sure sharing is correct. Error: {e}")

st.markdown("### 3) Download")
today = datetime.now().strftime("%Y%m%d")
default_name = f"walmart_price_update_{today}"
custom_name = st.text_input("Download file name (editable)", value=default_name)

hard_errors, not_found_skus, unpublished_skus, writable_out = validate_for_download(st.session_state.table_df)

with st.sidebar:
    st.divider()
    st.header("SKU Not Found on Walmart")
    if not_found_skus:
        st.text_area("Copy list", value="\n".join(map(str, not_found_skus)), height=220)
    else:
        st.caption("No SKUs in Not Found state.")

    st.divider()
    st.header("Unpublished SKU")
    if unpublished_skus:
        st.text_area("Copy list", value="\n".join(map(str, unpublished_skus)), height=220)
    else:
        st.caption("No Unpublished SKUs.")

if hard_errors:
    st.error("Hard Fail. Fix these issues before downloading:")
    for e in hard_errors:
        st.write(f"- {e}")
else:
    if unpublished_skus:
        st.warning(f"{len(unpublished_skus)} SKU are Unpublished. Do you still want to proceed?")
    else:
        st.success("No issues found. Ready to download.")

proceed_unpublished = False
if (not hard_errors) and unpublished_skus:
    proceed_unpublished = st.checkbox("Proceed even if SKU are Unpublished (include them in file)")

download_ready = False
download_bytes = None
download_filename = None

if TEMPLATE_PATH.exists() and (not hard_errors) and writable_out is not None and not writable_out.empty:
    if (not unpublished_skus) or (unpublished_skus and proceed_unpublished):
        download_bytes = fill_price_template(writable_out)
        safe = sanitize_filename(custom_name) or default_name
        download_filename = safe if safe.lower().endswith(".xlsx") else f"{safe}.xlsx"
        download_ready = True

st.download_button(
    label="Download Walmart Price Update File",
    data=download_bytes if download_ready else b"",
    file_name=download_filename if download_ready else "walmart_price_update.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    disabled=not download_ready,
)
